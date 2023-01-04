from add_archive import create_archive
from zipfile import ZipFile
import csv
import os
from PyPDF2 import PdfReader
from openpyxl import load_workbook


create_archive()
resources_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'resources')
archive_dir = os.path.join(resources_dir, 'testzip.zip')


def test_csv_file():
    with ZipFile(archive_dir) as zippfile:
        csvfile_dir = zippfile.extract(zippfile.namelist()[2])
        with open(csvfile_dir) as csvfile:
            csvfile = csv.reader(csvfile)
            csvread = []
            for r in csvfile:
                csvread.append(r)
            check_value = csvread[3][0].split(';')[0]
            assert check_value == 'johnson81'


def test_pdf_file():
    with ZipFile(archive_dir) as zippfile:
        pdffile_dir = zippfile.extract(zippfile.namelist()[1])
        pdffile = PdfReader(pdffile_dir)
        page = pdffile.pages[2]
        text = page.extract_text()
        assert 'Howtorundoctests' in text


def test_xlsx_file():
    with ZipFile(archive_dir) as zippfile:
        xlsxfile_dir = zippfile.extract(zippfile.namelist()[0])
        workbook = load_workbook(xlsxfile_dir)
        sheet = workbook.active
        check_value = sheet.cell(row=37, column=3).value
        assert check_value == 'Hacker'


def test_image_file():
    with ZipFile(archive_dir) as zippfile:
        imagefile_dir = zippfile.extract(zippfile.namelist()[0])
        check_size = os.path.getsize(imagefile_dir)
        print(check_size)
        assert check_size == 7360
